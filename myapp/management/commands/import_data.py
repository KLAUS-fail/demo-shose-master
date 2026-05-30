


from django.core.management.base import BaseCommand
import openpyxl
from myapp.models import Tovar, Zakaz, User


class Command(BaseCommand):
    help = "Импорт данных из excel"

    def handle(self, *args, **kwargs):
        wb = openpyxl.load_workbook("import/Tovar.xlsx")
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]: 
                Tovar.objects.get_or_create(articul=row[0])

        wb = openpyxl.load_workbook("import/Пункт выдачи_import.xlsx")
        for row in ws.active.iter_rows(min_row=2, values_only=True):
            if row[0]: 
                
                PunktVidachi.objects.get_or_create(id=row[0])
         

        wb = openpyxl.load_workbook("import/Заказ_import.xlsx")
        for row in ws.active.iter_rows(min_row=2, values_only=True):
            if row[0]: 
                punktVidachi = PunktVidachi.objects.get(id=row[4])
                Zakaz.objects.get_or_create(id=row[0])
         

